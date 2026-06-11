import re
from datetime import datetime

from django.db import transaction
from openpyxl import load_workbook

from accounts.services.portal_scope import (
    get_students_for_teacher,
    get_subjects_for_teacher,
    resolve_student_by_identifier,
)
from attendance.models import AttendanceRecord

_HEADER_ALIASES = {
    'roll_number': {'roll number', 'roll no', 'roll no.', 'rollno', 'registration no', 'registration number'},
    'date': {'date', 'attendance date', 'class date'},
    'status': {'status', 'attendance', 'attendance status'},
    'subject': {'subject', 'subject code', 'subject name', 'course'},
}


def _normalize_header(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value).strip().lower())


def _map_headers(header_row):
    mapping = {}
    for index, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        for field, aliases in _HEADER_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = index
    return mapping


def _parse_date(value):
    if value is None or str(value).strip() == '':
        return None
    if hasattr(value, 'date'):
        return value.date()
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_status(value):
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ('present', 'p', 'yes', 'y', '1'):
        return AttendanceRecord.STATUS_PRESENT
    if text in ('absent', 'a', 'no', 'n', '0'):
        return AttendanceRecord.STATUS_ABSENT
    return None


def _resolve_subject(value, teacher):
    if value is None or str(value).strip() == '':
        return None, 'Subject is required.'

    text = str(value).strip()
    subjects = list(get_subjects_for_teacher(teacher))
    for subject in subjects:
        if subject.code.lower() == text.lower() or subject.name.lower() == text.lower():
            return subject, None
    for subject in subjects:
        if text.lower() in subject.name.lower() or text.lower() in subject.code.lower():
            return subject, None
    return None, f'Subject "{text}" is not in your assigned subjects.'


def parse_attendance_workbook(uploaded_file, teacher):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {
            'valid_rows': [],
            'errors': ['The Excel file is empty.'],
            'duplicates': [],
            'summary': {'total': 0, 'valid': 0, 'errors': 1, 'duplicates': 0},
        }

    column_map = _map_headers(header_row)
    missing = [field for field in ('roll_number', 'date', 'status', 'subject') if field not in column_map]
    if missing:
        return {
            'valid_rows': [],
            'errors': [
                'Missing required columns. Use: Roll Number, Date, Status, Subject (row 1 headers).',
            ],
            'duplicates': [],
            'summary': {'total': 0, 'valid': 0, 'errors': 1, 'duplicates': 0},
        }

    valid_rows = []
    errors = []
    duplicates = []
    seen_keys = set()
    total = 0

    teacher_subjects = {s.pk: s for s in get_subjects_for_teacher(teacher)}

    for row_number, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        total += 1
        roll_value = row[column_map['roll_number']]
        date_value = row[column_map['date']]
        status_value = row[column_map['status']]
        subject_value = row[column_map['subject']]

        subject, subject_error = _resolve_subject(subject_value, teacher)
        if subject_error:
            errors.append(f'Row {row_number}: {subject_error}')
            continue

        record_date = _parse_date(date_value)
        if not record_date:
            errors.append(f'Row {row_number}: Invalid or missing date.')
            continue

        status = _parse_status(status_value)
        if not status:
            errors.append(f'Row {row_number}: Status must be Present or Absent.')
            continue

        student = resolve_student_by_identifier(
            roll_value,
            semester_id=subject.semester_id,
            department_id=subject.department_id,
        )
        if not student:
            errors.append(f'Row {row_number}: Student "{roll_value}" not found in roster for this subject.')
            continue

        roster = get_students_for_teacher(teacher, subject=subject)
        if not roster.filter(pk=student.pk).exists():
            errors.append(f'Row {row_number}: Student "{roll_value}" is not in your assigned roster.')
            continue

        key = (student.pk, subject.pk, record_date.isoformat())
        if key in seen_keys:
            duplicates.append(f'Row {row_number}: Duplicate entry for {student.display_roll_number} on {record_date}.')
            continue

        existing = AttendanceRecord.objects.filter(
            student=student,
            subject=subject,
            date=record_date,
        ).exists()
        if existing:
            duplicates.append(
                f'Row {row_number}: Attendance already recorded for {student.display_roll_number} '
                f'on {record_date} ({subject.code}).'
            )
            continue

        seen_keys.add(key)
        valid_rows.append({
            'student_id': student.pk,
            'subject_id': subject.pk,
            'date': record_date.isoformat(),
            'status': status,
            'student_name': student.user.display_name,
            'roll_number': student.display_roll_number,
            'subject_code': subject.code,
        })

    return {
        'valid_rows': valid_rows,
        'errors': errors,
        'duplicates': duplicates,
        'summary': {
            'total': total,
            'valid': len(valid_rows),
            'errors': len(errors),
            'duplicates': len(duplicates),
        },
    }


def commit_attendance_records(valid_rows, teacher):
    records = []
    for row in valid_rows:
        records.append(AttendanceRecord(
            student_id=row['student_id'],
            subject_id=row['subject_id'],
            date=row['date'],
            status=row['status'],
            recorded_by=teacher,
        ))

    with transaction.atomic():
        AttendanceRecord.objects.bulk_create(records)

    return len(records)


def get_attendance_summary_for_student(student):
    from accounts.services.portal_scope import get_subjects_for_student

    subjects = list(get_subjects_for_student(student))
    summaries = []

    for subject in subjects:
        records = AttendanceRecord.objects.filter(
            student=student,
            subject=subject,
        ).order_by('-date')
        total = records.count()
        attended = records.filter(status=AttendanceRecord.STATUS_PRESENT).count()
        percentage = round((attended / total) * 100) if total else 0

        if percentage >= 75:
            health = 'green'
        elif percentage >= 60:
            health = 'yellow'
        else:
            health = 'red'

        summaries.append({
            'subject': subject,
            'total': total,
            'attended': attended,
            'percentage': percentage,
            'health': health,
            'records': list(records),
        })

    overall_total = sum(s['total'] for s in summaries)
    overall_attended = sum(s['attended'] for s in summaries)
    overall_percentage = round((overall_attended / overall_total) * 100) if overall_total else 0

    return {
        'subjects': summaries,
        'overall_percentage': overall_percentage,
        'overall_total': overall_total,
        'overall_attended': overall_attended,
    }
