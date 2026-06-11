import re

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import load_workbook

from accounts.models import ActivityLog, Student
from accounts.services.user_service import create_user_account

User = get_user_model()

_HEADER_ALIASES = {
    'first_name': {'first_name', 'firstname', 'first name', 'fname'},
    'last_name': {'last_name', 'lastname', 'last name', 'lname', 'surname'},
    'email': {'email', 'email address', 'e-mail', 'mail'},
}


def _normalize_header(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value).strip().lower())


def _map_headers(header_row):
    mapping = {}
    for index, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        for field, aliases in _HEADER_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = index
    return mapping


def create_student_account(
    *,
    first_name,
    last_name,
    email,
    year,
    semester,
    performed_by=None,
):
    user, temp_password = create_user_account(
        first_name=first_name,
        last_name=last_name,
        email=email,
        role='student',
        year=year,
        semester=semester,
        performed_by=performed_by,
    )
    student = Student.objects.get(user=user)
    return student, temp_password


def update_student_profile(
    student,
    *,
    first_name,
    last_name,
    email,
    year,
    semester,
    performed_by=None,
):
    email = email.strip().lower()
    user = student.user
    if User.objects.filter(email=email).exclude(pk=user.pk).exists():
        raise ValueError('A student with this email already exists.')

    with transaction.atomic():
        user.first_name = first_name.strip()
        user.last_name = last_name.strip()
        user.email = email
        user.save(update_fields=['first_name', 'last_name', 'email'])

        student.year = year
        student.semester = semester
        student.save(update_fields=['year', 'semester'])

        ActivityLog.objects.create(
            performed_by=performed_by,
            action=ActivityLog.ACTION_STUDENT_UPDATED,
            description=f'Updated student profile for {user.display_name}',
        )

    return student


def import_students_from_excel(uploaded_file, *, year, semester, performed_by=None):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {'created': 0, 'skipped': 0, 'errors': ['The Excel file is empty.']}

    column_map = _map_headers(header_row)
    missing = [field for field in ('first_name', 'last_name', 'email') if field not in column_map]
    if missing:
        return {
            'created': 0,
            'skipped': 0,
            'errors': [
                'Missing required columns. Use: first_name, last_name, email '
                '(row 1 headers).',
            ],
        }

    created = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        first_name = str(row[column_map['first_name']] or '').strip()
        last_name = str(row[column_map['last_name']] or '').strip()
        email = str(row[column_map['email']] or '').strip().lower()

        if not first_name or not last_name or not email:
            errors.append(f'Row {row_number}: first name, last name, and email are required.')
            skipped += 1
            continue

        try:
            create_student_account(
                first_name=first_name,
                last_name=last_name,
                email=email,
                year=year,
                semester=semester,
                performed_by=performed_by,
            )
            created += 1
        except ValueError as exc:
            errors.append(f'Row {row_number}: {exc}')
            skipped += 1

    return {'created': created, 'skipped': skipped, 'errors': errors}


def get_students_for_teacher(teacher):
    from accounts.models import TeacherSubjectAssignment

    semester_ids = TeacherSubjectAssignment.objects.filter(
        teacher=teacher,
    ).values_list('subject__semester_id', flat=True).distinct()

    if not semester_ids:
        return Student.objects.none()

    queryset = Student.objects.filter(
        semester_id__in=semester_ids,
        user__role='student',
    ).select_related('user', 'year', 'year__department', 'semester')

    if teacher.department_id:
        queryset = queryset.filter(year__department_id=teacher.department_id)

    return queryset.order_by('user__first_name', 'user__last_name')
